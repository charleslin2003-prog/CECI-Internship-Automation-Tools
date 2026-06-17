import os
import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import LineString
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.spatial import cKDTree
import rasterio

# =========================================================================
# 1. 專案路徑配置（輸出完全打包收納在 0617 目錄下）
# =========================================================================
SHP_HUMAN = r"修正後.shp"
SHP_AI = r"SIGN_0527_修正後_raster_nearest.shp"
SHP_ROI = r"ROI_V01.shp"


TIFF_dir = r"d:\your_project\kriging_output"
TIFF_DX = os.path.join(TIFF_dir, "dX_kriging_1m.tif")
TIFF_DY = os.path.join(TIFF_dir, "dY_kriging_1m.tif")

out_dir = r"d:\your_project\report_output"
OUT_XLSX = os.path.join(out_dir, "spatial_accuracy_report.xlsx")
OUT_SHP_LINES = os.path.join(out_dir, "offset_vectors.shp")

COL_ROI_ID = "id"

# 🛠️ 核心門檻參數設定
RMSE_THRESHOLD   = 0.5        # RMSE 重算門檻（公尺）
DEFAULT_TAIWAN_CRS = "EPSG:3826"

# 🔍 配對搜尋參數
SEARCH_R1        = 1.0        # 第一階段：最近點搜尋半徑（公尺），找到即配對不做扇形
SEARCH_R2        = 5.0        # 第二階段：扇形搜尋半徑（公尺），第一階段無結果時啟用
SECTOR_HALF_DEG  = 20.0       # 第二階段扇形半角（度），總扇形 = SECTOR_HALF_DEG × 2

os.makedirs(out_dir, exist_ok=True)

# =========================================================================
# 2. GIS 空間幾何載入與真實座標萃取 (💡 已加入 CRS 遺失防護網)
# =========================================================================
print("⏳ [STEP 1] 載入圖層並同步座標系統...")
gdf_human = gpd.read_file(SHP_HUMAN)
gdf_ai = gpd.read_file(SHP_AI)
gdf_roi = gpd.read_file(SHP_ROI)

# 💡 【核心修復】自動防護無 CRS / 無 .prj 檔狀況
if gdf_human.crs is None:
    print(f"⚠️ 提示：偵測到 '{os.path.basename(SHP_HUMAN)}' 遺失座標系定義，已自動補回: {DEFAULT_TAIWAN_CRS}")
    gdf_human = gdf_human.set_crs(DEFAULT_TAIWAN_CRS, allow_override=True)

if gdf_ai.crs is None:
    print(f"⚠️ 提示：偵測到 '{os.path.basename(SHP_AI)}' 遺失座標系定義，已自動補回: {DEFAULT_TAIWAN_CRS}")
    gdf_ai = gdf_ai.set_crs(DEFAULT_TAIWAN_CRS, allow_override=True)

if gdf_roi.crs is None:
    print(f"⚠️ 提示：偵測到 '{os.path.basename(SHP_ROI)}' 遺失座標系定義，已自動補回: {DEFAULT_TAIWAN_CRS}")
    gdf_roi = gdf_roi.set_crs(DEFAULT_TAIWAN_CRS, allow_override=True)

# 安全轉換座標系統
if gdf_ai.crs != gdf_human.crs:
    gdf_ai = gdf_ai.to_crs(gdf_human.crs)
if gdf_roi.crs != gdf_human.crs:
    gdf_roi = gdf_roi.to_crs(gdf_human.crs)

# 強制從 Geometry 本體萃取 True X, Y 座標
gdf_human["true_h_x"] = gdf_human.geometry.x
gdf_human["true_h_y"] = gdf_human.geometry.y
gdf_ai["true_a_x"] = gdf_ai.geometry.x
gdf_ai["true_a_y"] = gdf_ai.geometry.y

gdf_human["_h_idx"] = range(len(gdf_human))
gdf_ai["_a_idx"] = range(len(gdf_ai))

# =========================================================================
# 3. 核心依據：讀取克里金 Tiff 進行「有依據的空間預測跳躍」
# =========================================================================
print("⏳ [STEP 2] 依據克里金網格計算人工點的形變跳躍，指引精準搜尋...")
h_pts_coords = [(r["true_h_x"], r["true_h_y"]) for _, r in gdf_human.iterrows()]
dx_tif_values, dy_tif_values = [], []

if os.path.exists(TIFF_DX) and os.path.exists(TIFF_DY):
    with rasterio.open(TIFF_DX) as src_dx:
        for val in src_dx.sample(h_pts_coords):
            dx_tif_values.append(
                val[0] if (hasattr(val, '__iter__') or isinstance(val, (list, np.ndarray))) else np.nan)
    with rasterio.open(TIFF_DY) as src_dy:
        for val in src_dy.sample(h_pts_coords):
            dy_tif_values.append(
                val[0] if (hasattr(val, '__iter__') or isinstance(val, (list, np.ndarray))) else np.nan)
else:
    # 💡 修正對象：精準對齊您的新變數名稱 TIFF_DX，提供正確的偵錯導引
    raise FileNotFoundError(f"❌ 錯誤：找不到指定的克里金 TIFF 改正網格檔，請確認路徑：\n{TIFF_DX}")

gdf_human["dX_tif"] = dx_tif_values
gdf_human["dY_tif"] = dy_tif_values

gdf_human["dX_tif"] = gdf_human["dX_tif"].fillna(0)
gdf_human["dY_tif"] = gdf_human["dY_tif"].fillna(0)

gdf_human["pred_h_x"] = gdf_human["true_h_x"] + gdf_human["dX_tif"]
gdf_human["pred_h_y"] = gdf_human["true_h_y"] + gdf_human["dY_tif"]

# =========================================================================
# 4. 以「最近點優先 + 扇形退回」進行 1對1 唯一配對
# =========================================================================
print("⏳ [STEP 3] 執行配對：1m 內取最近點；無結果則退回克里金方向扇形搜尋...")

SECTOR_HALF_RAD = np.radians(SECTOR_HALF_DEG)

coords_pred_human = np.column_stack((gdf_human["pred_h_x"], gdf_human["pred_h_y"]))
coords_ai         = np.column_stack((gdf_ai["true_a_x"],    gdf_ai["true_a_y"]))

# 從克里金 dX/dY 計算每個人工點的預測方向（弧度）
pred_angles   = np.arctan2(gdf_human["dY_tif"].values, gdf_human["dX_tif"].values)
has_direction = ~((gdf_human["dX_tif"].values == 0) & (gdf_human["dY_tif"].values == 0))

tree_ai        = cKDTree(coords_ai)
tree_pred      = cKDTree(coords_pred_human)

# 第一階段：SEARCH_R1 內所有候選（ai → pred_human）
stage1_indices = tree_ai.query_ball_tree(tree_pred, r=SEARCH_R1)
# 第二階段：SEARCH_R2 內所有候選，供扇形篩選用
stage2_indices = tree_ai.query_ball_tree(tree_pred, r=SEARCH_R2)

all_possible_matches = []

for ai_idx in range(len(coords_ai)):
    r1_human = stage1_indices[ai_idx]

    if r1_human:
        # ── 第一階段：1m 內有候選，直接加入（不做扇形）──────────────────
        for h_idx in r1_human:
            dist = np.linalg.norm(coords_pred_human[h_idx] - coords_ai[ai_idx])
            all_possible_matches.append({
                'h_idx': h_idx, 'a_idx': ai_idx,
                'dist_res': dist, 'stage': 1,
            })
    else:
        # ── 第二階段：1m 內無候選，退回扇形搜尋 ─────────────────────────
        for h_idx in stage2_indices[ai_idx]:
            # nodata 點（dX=dY=0）跳過扇形，直接納入
            if has_direction[h_idx]:
                vec        = coords_ai[ai_idx] - coords_pred_human[h_idx]
                angle_to_ai = np.arctan2(vec[1], vec[0])
                angle_diff  = (angle_to_ai - pred_angles[h_idx] + np.pi) % (2 * np.pi) - np.pi
                if abs(angle_diff) > SECTOR_HALF_RAD:
                    continue
            dist = np.linalg.norm(coords_pred_human[h_idx] - coords_ai[ai_idx])
            all_possible_matches.append({
                'h_idx': h_idx, 'a_idx': ai_idx,
                'dist_res': dist, 'stage': 2,
            })

df_matches_pool = pd.DataFrame(all_possible_matches)
valid_pair_rows = []

if not df_matches_pool.empty:
    df_matches_pool = df_matches_pool.sort_values(by="dist_res")
    used_human, used_ai = set(), set()
    for _, match in df_matches_pool.iterrows():
        h_id, a_id = int(match['h_idx']), int(match['a_idx'])
        if h_id not in used_human and a_id not in used_ai:
            used_human.add(h_id)
            used_ai.add(a_id)
            valid_pair_rows.append({'_h_idx': h_id, '_a_idx': a_id, 'stage': int(match['stage'])})

if not valid_pair_rows:
    raise ValueError("❌ 錯誤：兩階段搜尋均無結果，請檢查搜尋參數或圖層座標系是否一致。")

matched_human = gdf_human.merge(pd.DataFrame(valid_pair_rows), on="_h_idx", how="inner")
gdf_valid_pairs = matched_human.merge(gdf_ai.drop(columns="geometry"), on="_a_idx", how="inner")
# join 後去重，保留第一個匹配的 ROI，sjoin 前先重命名 ROI 欄位，避免衝突
gdf_roi_join = gdf_roi[[COL_ROI_ID, 'geometry']].rename(columns={COL_ROI_ID: 'roi_id'})
gdf_valid_pairs = gpd.sjoin(gdf_valid_pairs, gdf_roi_join,
                             how="left", predicate="intersects")
gdf_valid_pairs = gdf_valid_pairs[~gdf_valid_pairs.index.duplicated(keep='first')]

# 後續全改用 'roi_id'
COL_ROI_JOIN = 'roi_id'

# =========================================================================
# 5. 分區純幾何統計摘要與判定
# =========================================================================
print("⏳ [STEP 4] 計算真實二維偏移量與分區平面 RMSE 統計...")
gdf_valid_pairs["dX_obs"] = gdf_valid_pairs["true_a_x"] - gdf_valid_pairs["true_h_x"]
gdf_valid_pairs["dY_obs"] = gdf_valid_pairs["true_a_y"] - gdf_valid_pairs["true_h_y"]
gdf_valid_pairs["dist2d_obs"] = np.sqrt(gdf_valid_pairs["dX_obs"] ** 2 + gdf_valid_pairs["dY_obs"] ** 2)


def rmse(series):
    return np.sqrt(np.mean(series ** 2)) if len(series) > 0 else np.nan

def judge_pure_geometry(row):
    if pd.isna(row["rmse_2d"]):
        return "N/A (無配對點)"
    if float(row["rmse_2d"]) > RMSE_THRESHOLD:
        return "REDO (需重算)"
    return "PASS (合格)"


roi_stats_raw = gdf_valid_pairs.groupby(COL_ROI_JOIN).agg(
    pt_count=('dX_obs', 'count'),
    mean_2d=('dist2d_obs', 'mean'),
    rmse_2d=('dist2d_obs', rmse)
).reset_index()

# ── 補回無配對點的 ROI，避免從報表消失 ──────────────────────────────
all_roi_ids = gdf_roi[[COL_ROI_ID]].rename(columns={COL_ROI_ID: COL_ROI_JOIN}).drop_duplicates()
roi_stats_raw = all_roi_ids.merge(roi_stats_raw, on=COL_ROI_JOIN, how="left")
roi_stats_raw["pt_count"] = roi_stats_raw["pt_count"].fillna(0).astype(int)
# mean_2d / rmse_2d 保持 NaN，judge 會判為 N/A (無配對點)
# ─────────────────────────────────────────────────────────────────────

roi_stats_raw["Status"] = roi_stats_raw.apply(judge_pure_geometry, axis=1)

roi_stats = roi_stats_raw[[COL_ROI_JOIN, "pt_count", "mean_2d", "rmse_2d", "Status"]].copy()
df_redo = roi_stats[roi_stats["Status"] == "REDO (需重算)"].copy()

# =========================================================================
# 6. 匯出平面「偏移向量線」Shapefile
# =========================================================================
print("⏳ [STEP 5] 匯出 1對1 偏移向量線 Shapefile...")
lines_geometry = [LineString([(r["true_h_x"], r["true_h_y"]), (r["true_a_x"], r["true_a_y"])]) for _, r in
                  gdf_valid_pairs.iterrows()]

gdf_lines = gpd.GeoDataFrame(
    gdf_valid_pairs[[COL_ROI_JOIN, "dX_obs", "dY_obs", "dist2d_obs"]],
    geometry=lines_geometry,
    crs=gdf_human.crs  # 💡 引用安全指派後的明確 CRS，解決丟失問題
)
gdf_lines = gdf_lines.rename(columns={"dist2d_obs": "dist2d"})
gdf_lines.to_file(OUT_SHP_LINES, encoding="utf-8")

# =========================================================================
# 7. 打造經理指定：極簡高階 Excel 成果報表（開門見山、首行標頭）
# =========================================================================
print("⏳ [STEP 6] 正在將純幾何品管指標寫入 Excel...")
wb = openpyxl.Workbook()

NAVY_DARK = "1F497D"
ZEBRA_FILL = "F2F5F9"
BORDER_GRAY = "D9D9D9"
RED_ALERT_BG = "FDE9D9"
RED_ALERT_TXT = "C00000"
GREEN_PASS_BG = "E2EFDA"
GREEN_PASS_TXT = "375623"

font_header = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Microsoft JhengHei", size=10)
font_bold = Font(name="Microsoft JhengHei", size=10, bold=True)

fill_header = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
fill_zebra = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
fill_redo = PatternFill(start_color=RED_ALERT_BG, end_color=RED_ALERT_BG, fill_type="solid")
fill_pass = PatternFill(start_color=GREEN_PASS_BG, end_color=GREEN_PASS_BG, fill_type="solid")

thin_side, double_side, thin_top_side = Side(border_style="thin", color=BORDER_GRAY), Side(border_style="double",
                                                                                           color="000000"), Side(
    border_style="thin", color="000000")
border_cell, border_total = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side), Border(
    top=thin_top_side, bottom=double_side)
align_center, align_left, align_right = Alignment(horizontal="center", vertical="center"), Alignment(horizontal="left",
                                                                                                     vertical="center"), Alignment(
    horizontal="right", vertical="center")

# ---- TAB 1: ROI STATS ----
ws_stats = wb.active
ws_stats.title = "ROI精度統計表"
ws_stats.views.sheetView[0].showGridLines = True
ws_stats.freeze_panes = "A2"

headers_stats = ["ROI分區ID (id)", "唯一配對點數", "幾何平均偏移 (m)", "幾何平面 RMSE (m)", "判定結果"]

# 寫入表頭
for col_idx, text in enumerate(headers_stats, 1):
    cell = ws_stats.cell(row=1, column=col_idx, value=text)
    cell.font = font_header;
    cell.fill = fill_header;
    cell.alignment = align_center

# 寫入數據明細
for row_idx, row_data in enumerate(roi_stats.itertuples(index=False), 2):
    status_str = str(row_data[-1])
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_stats.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_body
        cell.border = border_cell
        # 先設底色：REDO 整列紅底，其餘奇數列斑馬
        if "REDO" in status_str:
            cell.fill = fill_redo
        elif row_idx % 2 == 1:
            cell.fill = fill_zebra
        # 再處理欄位格式（判定欄字體會在下面覆蓋）
        if col_idx == 1:
            cell.alignment = align_center
        elif col_idx == 2:
            cell.alignment = align_right
            cell.number_format = "#,##0"
        elif 3 <= col_idx <= 4:
            cell.alignment = align_right
            cell.number_format = "0.000"
        elif col_idx == 5:
            cell.alignment = align_center
            if "REDO" in status_str:
                cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color=RED_ALERT_TXT)
            else:
                cell.fill = fill_pass  # PASS 只有判定欄用綠色
                cell.font = Font(name="Microsoft JhengHei", size=10, bold=True, color=GREEN_PASS_TXT)

# 底部統計合計列
total_row = len(roi_stats) + 2
ws_stats.cell(row=total_row, column=1, value="整體平均 (Average)").font = font_bold;
ws_stats.cell(row=total_row, column=1).alignment = align_left;
ws_stats.cell(row=total_row, column=1).border = border_total
ws_stats.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})").number_format = "#,##0";
ws_stats.cell(row=total_row, column=2).font = font_bold;
ws_stats.cell(row=total_row, column=2).alignment = align_right;
ws_stats.cell(row=total_row, column=2).border = border_total
for c in range(3, 5):
    col_letter = get_column_letter(c)
    cell = ws_stats.cell(row=total_row, column=c, value=f"=AVERAGE({col_letter}2:{col_letter}{total_row - 1})")
    cell.number_format = "0.000"
    cell.font = font_bold;
    cell.alignment = align_right;
    cell.border = border_total
ws_stats.cell(row=total_row, column=5).border = border_total

# ---- TAB 2: REDO LIST ----
ws_redo = wb.create_sheet(title="需重算缺陷清單")
ws_redo.views.sheetView[0].showGridLines = True

for col_idx, text in enumerate(headers_stats, 1):
    cell = ws_redo.cell(row=1, column=col_idx, value=text);
    cell.font = font_header;
    cell.fill = PatternFill(start_color="951F1F", end_color="951F1F", fill_type="solid");
    cell.alignment = align_center

for row_idx, row_data in enumerate(df_redo.itertuples(index=False), 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_redo.cell(row=row_idx, column=col_idx, value=val);
        cell.font = font_body;
        cell.border = border_cell
        if col_idx == 1:
            cell.alignment = align_center
        elif col_idx == 2:
            cell.alignment = align_right; cell.number_format = "#,##0"
        elif 3 <= col_idx <= 4:
            cell.alignment = align_right; cell.number_format = "0.000"
        elif col_idx == 5:
            cell.alignment = align_center; cell.fill = fill_redo; cell.font = Font(name="Microsoft JhengHei", size=10,
                                                                                   bold=True, color=RED_ALERT_TXT)

# 自動調適優化欄寬
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                line_len = sum(2 if ord(char) > 128 else 1 for char in str(cell.value))
                if line_len > max_len: max_len = line_len
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(OUT_XLSX)
print("🎉 [品質成果收納完畢] 經典 5 欄純幾何版腳本無誤執行完畢！")
print(f"📈 瘦身成果 Excel 報表路徑：{OUT_XLSX}")
print(f"🗺️ 1對1 幾何偏移向量線 SHP：{OUT_SHP_LINES}")